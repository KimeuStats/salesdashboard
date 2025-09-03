import streamlit as st 
import pandas as pd
import numpy as np
import plotly.graph_objs as go
import base64
import requests
from st_aggrid import AgGrid, GridOptionsBuilder, JsCode
import io
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule

# === PAGE CONFIG ===
st.set_page_config(layout="wide", page_title="Muthokinju Paints Sales Dashboard")

# === STYLES ===
st.markdown("""
    <style>
        /* Container */
        .main .block-container {
            max-width: 1400px;
            padding: 2rem 2rem;
            margin: auto;
        }

        /* Banner */
        .banner {
            width: 100%;
            background-color: #3FA0A3;
            padding: 3px 30px;
            display: flex;
            align-items: center;
            justify-content: center;
            margin-bottom: 20px;
        }
        .banner img {
            height: 52px;
            margin-right: 15px;
            border: 2px solid white;
            box-shadow: 0 0 5px rgba(255,255,255,0.7);
        }
        .banner h1 {
            color: white;
            font-size: 26px;
            font-weight: bold;
            margin: 0;
        }

        /* Table Header */
        .ag-theme-material .ag-header {
            background-color: #7b38d8 !important;
            color: white !important;
            font-weight: bold !important;
        }

        /* Center Dashboard View title */
        .dashboard-view-title {
            text-align: center;
            font-weight: bold;
            margin-bottom: 1rem;
            font-size: 1.3rem;
        }

        /* View Selector container */
        .view-selector {
            display: flex;
            justify-content: center;
            gap: 20px;
            margin-bottom: 30px;
            flex-wrap: nowrap;
        }

        /* Buttons styled as cards */
        .view-button {
            padding: 15px 30px;
            border: 2px solid #7b38d8;
            border-radius: 12px;
            background-color: white;
            color: #7b38d8;
            font-weight: 700;
            cursor: pointer;
            box-shadow: 0 3px 8px rgba(123, 56, 216, 0.2);
            transition: all 0.3s ease;
            min-width: 140px;
            text-align: center;
            user-select: none;
        }
        .view-button:hover {
            background-color: #7b38d8;
            color: white;
            box-shadow: 0 5px 15px rgba(123, 56, 216, 0.4);
        }
        .view-button.active {
            background-color: #7b38d8;
            color: white;
            box-shadow: 0 5px 15px rgba(123, 56, 216, 0.6);
        }

        /* Responsive: On smaller screens (mobile), stack buttons in one horizontal scrollable row */
        @media (max-width: 600px) {
            .view-selector {
                justify-content: flex-start;
                gap: 12px;
                overflow-x: auto;
                padding-left: 10px;
            }
            .view-button {
                min-width: 120px;
                padding: 12px 18px;
                font-size: 0.9rem;
                flex-shrink: 0;
            }
        }
    </style>
""", unsafe_allow_html=True)

# === LOGO ===
# === Function to Load Base64 Image from Private GitHub ===
def load_base64_image_from_private_repo(owner, repo, path, token, branch="main"):
    url = f"https://api.github.com/repos/{owner}/{repo}/contents/{path}?ref={branch}"
    headers = {
        "Authorization": f"token {token}",
        "Accept": "application/vnd.github.v3.raw"
    }
    response = requests.get(url, headers=headers)
    if response.status_code == 200:
        return base64.b64encode(response.content).decode()
    else:
        st.error(f"⚠️ Failed to load image. GitHub API returned status {response.status_code}")
        return None

# === Load the Token from Streamlit Secrets ===
github_token = st.secrets["GITHUB_PAT"]

# === Parameters for Your Private Repo ===
github_owner = "kimeustats"
github_repo = "salesdashboard"
file_path = "nhmllogo.png"  # Path to image in repo
branch = "main"

# === Load and Display the Logo ===
logo_base64 = load_base64_image_from_private_repo(
    owner=github_owner,
    repo=github_repo,
    path=file_path,
    token=github_token,
    branch=branch
)

if logo_base64:
    st.markdown(f"""
        <div style="text-align:center;">
            <img src="data:image/png;base64,{logo_base64}" alt="Logo" width="200"/>
            <h1>Muthokinju Paints Sales Dashboard</h1>
        </div>
    """, unsafe_allow_html=True)
else:
    st.error("⚠️ Could not load logo.")
# === VIEW SELECTOR ===
st.markdown('<div class="dashboard-view-title">🧭 Dashboard View</div>', unsafe_allow_html=True)

# Wrap buttons in a div with view-selector class for flex styling
st.markdown('<div class="view-selector">', unsafe_allow_html=True)
view_col1, view_col2 = st.columns([1,1])
with view_col1:
    branch_view = st.button("🏢 Detailed View", key="branch_view", use_container_width=True)
with view_col2:
    general_view = st.button("🌐 General View", key="general_view", use_container_width=True)
st.markdown('</div>', unsafe_allow_html=True)

# Initialize session state for view
if 'current_view' not in st.session_state:
    st.session_state.current_view = 'branch'

if branch_view:
    st.session_state.current_view = 'branch'
elif general_view:
    st.session_state.current_view = 'general'

# Display current view with custom styling using markdown and CSS
active_class_branch = "view-button active" if st.session_state.current_view == 'branch' else "view-button"
active_class_general = "view-button active" if st.session_state.current_view == 'general' else "view-button"

# To visually reflect the active state, you can alternatively replace buttons by clickable divs, 
# but Streamlit buttons are a bit limited to fully style here. 
# So keep the buttons and add a markdown showing current view nicely:

current_view_display = "🏢 Detailed View" if st.session_state.current_view == 'branch' else "🌐 General View"
st.markdown(f"<p style='text-align:center; font-weight:bold; margin-top:10px;'>Current View: {current_view_display}</p>", unsafe_allow_html=True)

# === LOAD DATA ===
from io import BytesIO

# === Function to Load File from Private GitHub ===
def load_file_from_private_repo(owner, repo, path, token, branch="main"):
    url = f"https://api.github.com/repos/{owner}/{repo}/contents/{path}?ref={branch}"
    headers = {
        "Authorization": f"token {token}",
        "Accept": "application/vnd.github.v3.raw"
    }
    response = requests.get(url, headers=headers)
    if response.status_code == 200:
        return BytesIO(response.content)
    else:
        st.error(f"⚠️ Failed to load Excel file. GitHub API status: {response.status_code}")
        return None

# === Excel File Parameters ===
excel_path = "data1.xlsx"  # Make sure this matches exactly (case-sensitive)

# === Load Excel File Securely ===
excel_file = load_file_from_private_repo(
    owner=github_owner,
    repo=github_repo,
    path=excel_path,
    token=github_token,
    branch=branch
)

# DEBUG: Show constructed GitHub URL to verify
st.write(f"🔗 Fetching Excel from GitHub path: `{excel_path}` (branch: `{branch}`)")

# === Read DataFrames from Excel ===
try:
    if excel_file:
        sales = pd.read_excel(excel_file, sheet_name="CY", engine="openpyxl")
        st.success("✅ 'CY' sheet loaded successfully.")
        
        excel_file.seek(0)
        targets = pd.read_excel(excel_file, sheet_name="TARGETS", engine="openpyxl")
        st.success("✅ 'TARGETS' sheet loaded successfully.")
        
        excel_file.seek(0)
        prev_year_sales = pd.read_excel(excel_file, sheet_name="PY", engine="openpyxl")
        st.success("✅ 'PY' sheet loaded successfully.")
    else:
        st.error("❌ Excel file is `None`. It was not fetched correctly from GitHub.")
        st.stop()
except Exception as e:
    st.error(f"⚠️ Failed to load Excel data: {e}")
    st.stop()



# === CLEAN DATA ===
sales.columns = [col if col == 'Cluster' else col.lower() for col in sales.columns]
targets.columns = targets.columns.str.lower()
prev_year_sales.columns = prev_year_sales.columns.str.lower()

sales['date'] = pd.to_datetime(sales['date'])
prev_year_sales['date'] = pd.to_datetime(prev_year_sales['date'])

for df in [sales, targets, prev_year_sales]:
    df['amount'] = df['amount'].astype(str).str.replace(',', '').astype(float)

targets_agg = targets.groupby(['branch', 'category1'], as_index=False)['amount'].sum().rename(columns={'amount': 'monthly_target'})

# === HELPER FUNCTION ===
def working_days_excl_sundays(start_date, end_date):
    return len([d for d in pd.date_range(start=start_date, end=end_date) if d.weekday() != 6])

# === FILTERS ===
clusters = sales["Cluster"].dropna().unique()
branches = sales["branch"].dropna().unique()
categories = sales["category1"].dropna().unique()
date_min, date_max = sales["date"].min(), sales["date"].max()

# Dynamic filters based on view
if st.session_state.current_view == 'branch':
    # Branch View - show all filters including branch
    col1, col2, col3 = st.columns(3)
    with col1:
        selected_cluster = st.selectbox("Cluster", options=["All"] + list(clusters))
    with col2:
        selected_branch = st.selectbox("Branch", options=["All"] + list(branches))
    with col3:
        selected_category = st.selectbox("Category", options=["All"] + list(categories))
else:
    # General View - no branch filter, cluster shows per category
    col1, col2 = st.columns(2)
    with col1:
        selected_cluster = st.selectbox("Cluster", options=["All"] + list(clusters))
    with col2:
        selected_category = st.selectbox("Category", options=["All"] + list(categories))
    selected_branch = "All"  # Always set to All for general view

col_from, col_to = st.columns(2)
with col_from:
    st.markdown("<div style='background-color:#7b38d8; color:white; padding:8px; font-weight:bold;'>From</div>", unsafe_allow_html=True)
    start_date = st.date_input("", value=date_min, min_value=date_min, max_value=date_max, key="from_date")
with col_to:
    st.markdown("<div style='background-color:#7b38d8; color:white; padding:8px; font-weight:bold;'>To</div>", unsafe_allow_html=True)
    end_date = st.date_input("", value=date_max, min_value=date_min, max_value=date_max, key="to_date")

# === APPLY FILTERS ===
filtered = sales.copy()
if selected_cluster != "All":
    filtered = filtered[filtered["Cluster"] == selected_cluster]
if selected_branch != "All":
    filtered = filtered[filtered["branch"] == selected_branch]
if selected_category != "All":
    filtered = filtered[filtered["category1"] == selected_category]
filtered = filtered[(filtered["date"] >= pd.to_datetime(start_date)) & (filtered["date"] <= pd.to_datetime(end_date))]

if filtered.empty:
    st.warning("⚠️ No sales data found for the selected filters or date range.")
    st.stop()

# === CORRECT WORKING DAYS LOGIC ===
end_dt = pd.to_datetime(end_date)
month_start = pd.Timestamp(end_dt.year, end_dt.month, 1)
month_end = pd.Timestamp(end_dt.year, end_dt.month, end_dt.days_in_month)

days_worked = working_days_excl_sundays(month_start, end_dt)
total_working_days = working_days_excl_sundays(month_start, month_end)

# === AGGREGATIONS ===
if st.session_state.current_view == 'general':

    if selected_cluster == "All":
        # --- Aggregate across all clusters (sum by category only) ---
        mtd_agg = filtered.groupby(['category1'], as_index=False)['amount'].sum().rename(columns={'amount': 'mtd_achieved'})
        daily_achieved = (
            filtered[filtered['date'] == end_dt]
            .groupby(['category1'], as_index=False)['amount']
            .sum()
            .rename(columns={'amount': 'daily_achieved'})
        )

        # Targets: group by category only
        targets_general = (
            targets.groupby(['category1'], as_index=False)['amount']
            .sum()
            .rename(columns={'amount': 'monthly_target'})
        )

        # Previous year sales - ADJUSTED: MTD for current month date selection
        prev_year_start = pd.Timestamp(end_dt.year - 1, end_dt.month, 1)
        prev_year_end_dt = pd.Timestamp(end_dt.year - 1, end_dt.month, end_dt.day)
        
        prev_year_filtered = prev_year_sales[
            (prev_year_sales['date'] >= prev_year_start) &
            (prev_year_sales['date'] <= prev_year_end_dt)
        ]
        if selected_category != "All":
            prev_year_filtered = prev_year_filtered[prev_year_filtered["category1"] == selected_category]

        pym_agg = (
            prev_year_filtered
            .groupby(['category1'], as_index=False)['amount']
            .sum()
            .rename(columns={'amount': 'pym'})
        )

        # Merge all
        df = (
            mtd_agg
            .merge(daily_achieved, on='category1', how='left')
            .merge(targets_general, on='category1', how='left')
            .merge(pym_agg, on='category1', how='left')
        )
        
        df.fillna(0, inplace=True)
        df.insert(0, 'branch', 'All Clusters')  # First column for compatibility

    else:
        # --- When a specific cluster is selected ---
        mtd_agg = (
            filtered.groupby(['Cluster', 'category1'], as_index=False)['amount']
            .sum()
            .rename(columns={'amount': 'mtd_achieved'})
        )
        daily_achieved = (
            filtered[filtered['date'] == end_dt]
            .groupby(['Cluster', 'category1'], as_index=False)['amount']
            .sum()
            .rename(columns={'amount': 'daily_achieved'})
        )

        # Targets
        targets_general = (
            targets.groupby(['cluster', 'category1'], as_index=False)['amount']
            .sum()
            .rename(columns={'amount': 'monthly_target', 'cluster': 'Cluster'})
        )

        # Previous year filtering - ADJUSTED: MTD for current month date selection
        prev_year_start = pd.Timestamp(end_dt.year - 1, end_dt.month, 1)
        prev_year_end_dt = pd.Timestamp(end_dt.year - 1, end_dt.month, end_dt.day)
        
        prev_year_filtered = prev_year_sales[
            (prev_year_sales['date'] >= prev_year_start) &
            (prev_year_sales['date'] <= prev_year_end_dt)
        ]
        prev_year_filtered = prev_year_filtered[prev_year_filtered["cluster"] == selected_cluster]
        if selected_category != "All":
            prev_year_filtered = prev_year_filtered[prev_year_filtered["category1"] == selected_category]

        pym_agg = (
            prev_year_filtered
            .groupby(['cluster', 'category1'], as_index=False)['amount']
            .sum()
            .rename(columns={'amount': 'pym', 'cluster': 'Cluster'})
        )

        # Merge all
        df = (
            mtd_agg
            .merge(daily_achieved, on=['Cluster', 'category1'], how='left')
            .merge(targets_general, on=['Cluster', 'category1'], how='left')
            .merge(pym_agg, on=['Cluster', 'category1'], how='left')
        )

        df.fillna(0, inplace=True)
        df = df.rename(columns={'Cluster': 'branch'})  # Ensure compatibility

else:
    # --- Branch view logic ---
    mtd_agg = (
        filtered.groupby(['branch', 'category1'], as_index=False)['amount']
        .sum()
        .rename(columns={'amount': 'mtd_achieved'})
    )
    daily_achieved = (
        filtered[filtered['date'] == end_dt]
        .groupby(['branch', 'category1'], as_index=False)['amount']
        .sum()
        .rename(columns={'amount': 'daily_achieved'})
    )

    # Previous year filtering - ADJUSTED: MTD for current month date selection
    prev_year_start = pd.Timestamp(end_dt.year - 1, end_dt.month, 1)
    prev_year_end_dt = pd.Timestamp(end_dt.year - 1, end_dt.month, end_dt.day)
    
    prev_year_filtered = prev_year_sales[
        (prev_year_sales['date'] >= prev_year_start) &
        (prev_year_sales['date'] <= prev_year_end_dt)
    ]

    pym_agg = (
        prev_year_filtered
        .groupby(['branch', 'category1'], as_index=False)['amount']
        .sum()
        .rename(columns={'amount': 'pym'})
    )

    df = (
        mtd_agg
        .merge(daily_achieved, on=['branch', 'category1'], how='left')
        .merge(targets_agg, on=['branch', 'category1'], how='left')
        .merge(pym_agg, on=['branch', 'category1'], how='left')
    )
    df.fillna(0, inplace=True)



# === CALCULATIONS ===
df['daily_tgt'] = np.where(total_working_days>0, df['monthly_target']/total_working_days, 0)
df['achieved_vs_daily_tgt'] = np.where(df['daily_tgt']>0, (df['daily_achieved'] - df['daily_tgt']) / df['daily_tgt'], 0)
df['mtd_tgt'] = df['daily_tgt'] * days_worked
df['mtd_var'] = np.where(df['mtd_tgt']>0, (df['mtd_achieved'] - df['mtd_tgt']) / df['mtd_tgt'], 0)
df['cm'] = df['mtd_achieved']
df['achieved_vs_monthly_tgt'] = np.where(df['monthly_target']>0, (df['mtd_achieved'] - df['monthly_target']) / df['monthly_target'], 0)
df['projected_landing'] = np.where(days_worked>0, (df['mtd_achieved'] / days_worked) * total_working_days, 0)
df['cm_vs_pym'] = np.where(df['pym']>0, (df['cm'] - df['pym']) / df['pym'], 0)

df.rename(columns={
    'monthly_target': 'Monthly TGT',
    'daily_tgt': 'Daily Tgt',
    'daily_achieved': 'Daily Achieved',
    'achieved_vs_daily_tgt': 'Achieved vs Daily Tgt',
    'mtd_tgt': 'MTD TGT',
    'mtd_achieved': 'MTD Act.',
    'mtd_var': 'MTD Var',
    'cm': 'CM',
    'achieved_vs_monthly_tgt': 'Achieved VS Monthly tgt',
    'projected_landing': 'Projected landing',
    'pym': 'PYM',
    'cm_vs_pym': 'CM VS PYM'
}, inplace=True)

# === KPI CALCULATIONS ===
current_view = st.session_state.current_view
current_branch = selected_branch if current_view == 'branch' else None

kpi1 = df['MTD Act.'].sum()

if current_view == 'branch':
    if current_branch and current_branch != "All":
        # Specific branch selected - use paints target for that branch only
        paints_rows = df[(df['category1'].str.lower() == 'paints') & (df['branch'] == current_branch)]
        kpi2 = paints_rows['Monthly TGT'].sum() if not paints_rows.empty else 0
    else:
        # All branches selected - sum paints targets across all branches from original targets data
        paints_targets = targets_agg[targets_agg['category1'].str.lower() == 'paints']
        # Apply cluster and category filters if any
        if selected_cluster != "All":
            # Filter targets by cluster - need to map branch to cluster
            cluster_branches = sales[sales["Cluster"] == selected_cluster]["branch"].unique()
            paints_targets = paints_targets[paints_targets['branch'].isin(cluster_branches)]
        kpi2 = paints_targets['monthly_target'].sum() if not paints_targets.empty else 0
else:
    paints_rows = df[df['category1'].str.lower() == 'paints']
    kpi2 = paints_rows['Monthly TGT'].sum() if not paints_rows.empty else 0

kpi3 = df['Daily Achieved'].sum()
kpi4 = df['Projected landing'].sum()

# === STYLES ===
st.markdown("""
<style>
.kpi-grid {
    display: flex;
    flex-wrap: wrap;
    gap: 16px;
    margin-top: 10px;
    justify-content: space-between;
}
.kpi-box {
    flex: 1 1 calc(20% - 16px);
    background-color: #f7f7fb;
    border-left: 6px solid #7b38d8;
    border-radius: 10px;
    padding: 16px;
    min-width: 150px;
    box-shadow: 1px 1px 4px rgba(0,0,0,0.05);
}
.kpi-box h4 {
    margin: 0;
    font-size: 14px;
    color: #555;
    font-weight: 600;
}
.kpi-box p {
    margin: 5px 0 0 0;
    font-size: 22px;
    font-weight: bold;
    color: #222;
}
@media only screen and (max-width: 768px) {
    .kpi-box {
        flex: 1 1 calc(48% - 16px);
    }
}
</style>
""", unsafe_allow_html=True)

# === KPI DISPLAY ===
st.markdown(f"""
<div class="kpi-grid">
    <div class="kpi-box">
        <h4>🏅 MTD Achieved</h4>
        <p>{kpi1:,.0f}</p>
    </div>
    <div class="kpi-box">
        <h4>🚀 Monthly Target</h4>
        <p>{kpi2:,.0f}</p>
    </div>
    <div class="kpi-box">
        <h4>📅 Daily Achieved</h4>
        <p>{kpi3:,.0f}</p>
    </div>
    <div class="kpi-box">
        <h4>📈 Projected Landing</h4>
        <p>{kpi4:,.0f}</p>
    </div>
    <div class="kpi-box">
        <h4>⌛ Days Worked</h4>
        <p>{days_worked} / {total_working_days}</p>
    </div>
</div>
""", unsafe_allow_html=True)

# === SALES VS TARGET CHART ===
chart_title = "📊 Sales vs Monthly Target (MTD)" + (" - General View" if st.session_state.current_view == 'general' else " - Detailed View")
st.markdown(f"### {chart_title}")
df_chart = df.copy()
x_labels = df_chart.apply(lambda row: f"{row['branch']} - {row['category1']}", axis=1)

fig = go.Figure([
    go.Bar(x=x_labels, y=df_chart['MTD Act.'], name='MTD Achieved', marker_color='orange'),
    go.Bar(x=x_labels, y=df_chart['Monthly TGT'], name='Monthly Target', marker_color='steelblue')
])
fig.update_layout(barmode='group', xaxis_tickangle=-45,
                  height=500, margin=dict(b=150),
                  legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1))
st.plotly_chart(fig, use_container_width=True)

# === AGGRID DISPLAY with Totals Row ===
df_display = df.copy()
percent_cols = ['Achieved vs Daily Tgt', 'MTD Var', 'Achieved VS Monthly tgt', 'CM VS PYM']

# FIXED TOTALS LOGIC: Use paints row for targets in both views, sum all achieved values
paints_row = df_display[df_display['category1'].str.lower() == 'paints']

if paints_row.empty:
    st.warning("⚠️ 'Paints' row not found — totals may be inaccurate.")
    paints_values = {col: 0 for col in ['Monthly TGT', 'Daily Tgt', 'MTD TGT', 'PYM']}
else:
    paints_values = {
        'Monthly TGT': paints_row['Monthly TGT'].values[0],
        'Daily Tgt': paints_row['Daily Tgt'].values[0],
        'MTD TGT': paints_row['MTD TGT'].values[0],
        'PYM': paints_row['PYM'].values[0]
    }

# Sum all achieved values (paints + other categories)
actual_sums = df_display[['Daily Achieved', 'MTD Act.', 'Projected landing', 'CM']].sum()

# Calculate percentages
def safe_div(n, d): return (n - d) / d if d else 0

totals = {
    'branch': 'Totals',
    'category1': '',
    'Monthly TGT': paints_values['Monthly TGT'],
    'Daily Tgt': paints_values['Daily Tgt'],
    'MTD TGT': paints_values['MTD TGT'],
    'PYM': paints_values['PYM'],
    'Daily Achieved': actual_sums['Daily Achieved'],
    'MTD Act.': actual_sums['MTD Act.'],
    'Projected landing': actual_sums['Projected landing'],
    'CM': actual_sums['CM'],
    'Achieved vs Daily Tgt': safe_div(actual_sums['Daily Achieved'], paints_values['Daily Tgt']),
    'MTD Var': safe_div(actual_sums['MTD Act.'], paints_values['MTD TGT']),
    'Achieved VS Monthly tgt': safe_div(actual_sums['MTD Act.'], paints_values['Monthly TGT']),
    'CM VS PYM': safe_div(actual_sums['CM'], paints_values['PYM']),
    'is_totals': True
}

# Append Totals row
df_display = pd.concat([df_display, pd.DataFrame([totals])], ignore_index=True)

# Formatting
for col in percent_cols:
    df_display[col] = (df_display[col].astype(float) * 100).round(1)
for col in df_display.columns:
    if pd.api.types.is_numeric_dtype(df_display[col]) and col not in percent_cols:
        df_display[col] = df_display[col].round(1)

# AgGrid setup
gb = GridOptionsBuilder.from_dataframe(df_display)
gb.configure_default_column(filter=True, sortable=True, resizable=True, autoHeight=True)
gb.configure_column("is_totals", hide=True)

# Style for % columns
cell_style_jscode = JsCode("""
function(params) {
    if (params.value == null) return {};
    if (params.value < 0) {
        return {color: 'black', backgroundColor: '#ffc0cb', fontWeight: 'bold', textAlign: 'center'};
    } else if (params.value > 0) {
        return {color: 'black', backgroundColor: '#d0f0c0', textAlign: 'center'};
    }
    return {textAlign: 'center'};
}
""")

# Apply formatting for % columns
for col in percent_cols:
    gb.configure_column(
        col,
        cellStyle=cell_style_jscode,
        type=["numericColumn", "numberColumnFilter", "customNumericFormat"],
        valueFormatter="x.toFixed(1) + '%'",
        headerClass='header-center'
    )

# Apply comma formatting to numeric (non-percentage) columns
for col in df_display.columns:
    if pd.api.types.is_numeric_dtype(df_display[col]) and col not in percent_cols:
        gb.configure_column(
            col,
            type=["numericColumn", "numberColumnFilter", "customNumericFormat"],
            valueFormatter=JsCode("""
                function(params) {
                    return params.value != null 
                        ? params.value.toLocaleString(undefined, {minimumFractionDigits: 1, maximumFractionDigits: 1}) 
                        : '';
                }
            """),
            headerClass='header-center'
        )

# Totals row styling
gb.configure_grid_options(getRowStyle=JsCode("""
function(params) {
    if (params.data.is_totals) {
        return {
            backgroundColor: '#b2dfdb',
            fontWeight: 'bold',
            fontSize: '14px',
            textAlign: 'center'
        };
    }
    return {};
}
"""))

st.markdown("<style>.ag-theme-material .ag-cell{text-align:center !important;}</style>", unsafe_allow_html=True)

table_title = f"### <center>📋 <span style='font-size:22px; font-weight:bold; color:#7b38d8;'>PERFORMANCE TABLE - {current_view_display}</span></center>"
st.markdown(table_title, unsafe_allow_html=True)
AgGrid(df_display, gridOptions=gb.build(), enable_enterprise_modules=False,
       allow_unsafe_jscode=True, theme="material", height=500, fit_columns_on_grid_load=False, reload_data=True)

# === EXCEL DOWNLOAD ===
df_excel = df_display.drop(columns=['is_totals', '::auto_unique_id::'], errors='ignore').copy()
for col in percent_cols:
    df_excel[col] = df_excel[col] / 100  # revert to decimal for Excel

excel_buffer = io.BytesIO()
with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
    df_excel.to_excel(writer, index=False, sheet_name='Performance')
    ws = writer.sheets['Performance']
    header = list(df_excel.columns)
    fill_neg = PatternFill(start_color='FFC0CB', end_color='FFC0CB', fill_type='solid')
    fill_pos = PatternFill(start_color='D0F0C0', end_color='D0F0C0', fill_type='solid')
    for col_name in percent_cols:
        if col_name in header:
            col_idx = header.index(col_name) + 1
            for row in range(2, len(df_excel) + 2):
                ws.cell(row=row, column=col_idx).number_format = '0.0%'
            ws.conditional_formatting.add(
                f"{openpyxl.utils.get_column_letter(col_idx)}2:{openpyxl.utils.get_column_letter(col_idx)}{len(df_excel)+1}",
                CellIsRule(operator='lessThan', formula=['0'], fill=fill_neg))
            ws.conditional_formatting.add(
                f"{openpyxl.utils.get_column_letter(col_idx)}2:{openpyxl.utils.get_column_letter(col_idx)}{len(df_excel)+1}",
                CellIsRule(operator='greaterThan', formula=['0'], fill=fill_pos))

excel_buffer.seek(0)

view_suffix = "_general_view" if st.session_state.current_view == 'general' else "_branch_view"
filename = f"sales_dashboard{view_suffix}.xlsx"

st.download_button(label="📥 Download Table as Excel",
                   data=excel_buffer,
                   file_name=filename,
                   mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
